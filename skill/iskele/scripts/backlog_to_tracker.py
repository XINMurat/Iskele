#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
backlog_to_tracker.py — backlog markdown'ini takip cizelgesine (xlsx) cevirir.

Beklenen backlog formati (iskele/assets/templates/03-gorev-listesi.md):

    ### Epik F1.2 — Surec & adim
    - [ ] **F1-BE-04** (M) Surec CRUD API. **Bag.:** F1-BE-02, F1-BE-03
      - *Kabul:* Yetki kontrolüyle calisir.

Uretilen sekmeler: Takip (gorev satirlari), Ozet (COUNTIF formulleri), Aciklama.
"Durum/Sorumlu/Baslangic/Bitis/Not" sutunlari elle doldurulur.

Kullanim:
    python backlog_to_tracker.py --backlog 03-gorev-listesi.md --out tracker.xlsx
    python backlog_to_tracker.py ... --config iskele.config.json
"""
import argparse, json, re, sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
# "Hakem" SONA eklendi, araya degil: Durum H sutununda kalmali — Ozet
# sekmesindeki COUNTIF/COUNTIFS formulleri ve acilir liste dogrulamasi ona
# bagli. Sutun sirasini estetik gerekceyle degistirmek o formulleri sessizce
# yanlis hucreye baglar.
HEADERS = ["ID", "Faz", "Epik", "Gorev", "Katman", "Tahmin",
           "Bagimlilik", "Durum", "Sorumlu", "Baslangic", "Bitis", "Not",
           "Hakem"]
STATUSES = ["Yapilacak", "Devam", "Bloke", "Tamamlandi"]
PALETTE = ["E2EFDA", "DEEBF7", "FCE4D6", "EDEDED", "FFF2CC", "E4DFEC"]

EPIC_RE = re.compile(r'^#{2,4}\s*Epik\s+(?P<code>[A-Za-z0-9.\-]+)\s*[—\-–:]?\s*(?P<name>.*)$')
TASK_RE = re.compile(
    r'^\s*[-*]\s*\[[ x~]\]\s*\*\*(?P<id>[A-Za-z0-9.\-_]+)\*\*\s*'
    r'\((?P<est>[SMLsml])\)\s*(?P<rest>.+)$')
DEP_RE = re.compile(r'\*\*(?:Bag\.?|Bağ\.?|Dep\.?|Bagimlilik|Bağımlılık)\s*:?\*\*\s*(?P<deps>.*)$',
                    re.IGNORECASE)
# Kabul kriteri gorev satirinin ALTINDA, girintili bir alt-madde olarak durur.
# Cizelgeye tasinmaz (orada sutunu yok) ama iskele_to_registry.py bunu
# onkayitli curutme kosuluna cevirir; o yuzden parser burada yakalar.
KABUL_RE = re.compile(r'^\s+[-*]\s*\*?(?:Kabul|Acceptance)\*?\s*:?\*?\s*(?P<kabul>.+?)\*?\s*$',
                      re.IGNORECASE)
# Kabul kriterinin icindeki opsiyonel hakem: "... **Hakem:** pytest tests/x.py".
# Yazilmadiginda BOS kalir — varsayilan uydurulmaz; bos hucre "bu gorevin
# 'Tamamlandi'si oz-beyandir" demektir ve rapor bunu oyle basar.
HAKEM_RE = re.compile(r'\*\*(?:Hakem|Arbiter)\*?\s*:?\*\*\s*(?P<hakem>.+?)\s*$',
                      re.IGNORECASE)


def parse_backlog(path):
    """markdown -> (tasks, problems)"""
    tasks, problems, seen = [], [], {}
    epic_code = epic_name = ""
    in_fence = False
    fence_marker = None
    for lineno, raw in enumerate(Path(path).read_text(encoding="utf-8").splitlines(), 1):
        # Cit icindeki satirlar ORNEKTIR, gorev degil. Backlog'lar bicim ornegi
        # tasir ("- [ ] **F1-BE-01** (M) ...") ve cit atlanmazsa o ornek canli bir
        # gorev olarak cizelgeye girer: sayim ve efor sessizce sisar. Acilis
        # isaretini sakla, ayni uzunluktaki kapanis kapatsin -- ic ice ``` bloklari
        # erken kapanmasin.
        stripped = raw.lstrip()
        if stripped.startswith("```") or stripped.startswith("~~~"):
            marker = stripped[:3]
            if not in_fence:
                in_fence, fence_marker = True, marker
            elif marker == fence_marker:
                in_fence, fence_marker = False, None
            continue
        if in_fence:
            continue
        m = EPIC_RE.match(raw.strip())
        if m:
            epic_code = m.group("code").strip()
            epic_name = m.group("name").strip()
            continue
        m = TASK_RE.match(raw)
        if not m:
            km = KABUL_RE.match(raw)
            if km and tasks:
                kabul = km.group("kabul").strip()
                tasks[-1]["kabul"] = kabul
                hm = HAKEM_RE.search(kabul)
                if hm:
                    tasks[-1]["hakem"] = hm.group("hakem").strip()
            continue
        tid = m.group("id").strip()
        if tid in seen:
            problems.append(f"satir {lineno}: '{tid}' ID'si mukerrer (ilk: satir {seen[tid]})")
        seen[tid] = lineno

        rest = m.group("rest").strip()
        deps = ""
        dm = DEP_RE.search(rest)
        if dm:
            deps = dm.group("deps").strip().rstrip(".")
            rest = rest[:dm.start()].strip()
        title = rest.rstrip(". ").strip()
        if not title:
            problems.append(f"satir {lineno} [{tid}]: gorev basligi bos")

        parts = tid.split("-")
        faz = parts[0] if parts else ""
        katman = parts[1] if len(parts) > 2 else ""

        if not epic_code:
            problems.append(f"satir {lineno} [{tid}]: epik basligi yok "
                            f"('### Epik <kod> — <ad>' ekle)")

        tasks.append(dict(id=tid, faz=faz,
                          epik=f"{epic_code} {epic_name}".strip(),
                          gorev=title, katman=katman,
                          tahmin=m.group("est").upper(), bagimlilik=deps,
                          kabul="", hakem="", satir=lineno))
    if not tasks:
        problems.append("hic gorev satiri bulunamadi — backlog formatini kontrol et")
    return tasks, problems


def build(tasks, out_path, phases):
    wb = Workbook()
    ws = wb.active
    ws.title = "Takip"

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    fills = {p: PatternFill("solid", fgColor=PALETTE[i % len(PALETTE)])
             for i, p in enumerate(phases)}

    for r, t in enumerate(tasks, start=2):
        values = [t["id"], t["faz"], t["epik"], t["gorev"], t["katman"],
                  t["tahmin"], t["bagimlilik"], "Yapilacak", "", "", "", "",
                  t.get("hakem", "")]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (3, 4, 7, 12)))
            if c == 2 and t["faz"] in fills:
                cell.fill = fills[t["faz"]]

    last = len(tasks) + 1
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUSES), allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"H2:H{last}")

    for i, w in enumerate([12, 6, 26, 42, 8, 8, 24, 13, 14, 13, 13, 30, 34], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{last}"
    ws.sheet_view.showGridLines = False

    # ---- Ozet
    wo = wb.create_sheet("Ozet")
    wo.sheet_view.showGridLines = False
    wo.cell(row=1, column=1, value="Ilerleme Ozeti").font = Font(
        name=FONT, bold=True, size=14, color="1F3864")

    wo.cell(row=3, column=1, value="Durum").font = Font(name=FONT, bold=True)
    wo.cell(row=3, column=2, value="Adet").font = Font(name=FONT, bold=True)
    for i, s in enumerate(STATUSES):
        row = 4 + i
        wo.cell(row=row, column=1, value=s).font = Font(name=FONT)
        wo.cell(row=row, column=2,
                value=f'=COUNTIF(Takip!$H$2:$H${last},A{row})').font = Font(name=FONT)
    done_row = 4 + STATUSES.index("Tamamlandi")
    tot_row = 4 + len(STATUSES)
    wo.cell(row=tot_row, column=1, value="Toplam").font = Font(name=FONT, bold=True)
    wo.cell(row=tot_row, column=2,
            value=f'=COUNTA(Takip!$A$2:$A${last})').font = Font(name=FONT, bold=True)
    wo.cell(row=tot_row + 1, column=1, value="Tamamlanma %").font = Font(name=FONT, bold=True)
    pc = wo.cell(row=tot_row + 1, column=2,
                 value=f'=IF(B{tot_row}=0,0,B{done_row}/B{tot_row})')
    pc.font = Font(name=FONT, bold=True)
    pc.number_format = "0.0%"

    for j, h in enumerate(["Faz", "Toplam", "Tamamlanan", "%"], start=4):
        wo.cell(row=3, column=j, value=h).font = Font(name=FONT, bold=True)
    for i, p in enumerate(phases):
        row = 4 + i
        wo.cell(row=row, column=4, value=p).font = Font(name=FONT)
        wo.cell(row=row, column=5,
                value=f'=COUNTIF(Takip!$B$2:$B${last},D{row})').font = Font(name=FONT)
        wo.cell(row=row, column=6,
                value=f'=COUNTIFS(Takip!$B$2:$B${last},D{row},'
                      f'Takip!$H$2:$H${last},"Tamamlandi")').font = Font(name=FONT)
        cell = wo.cell(row=row, column=7, value=f'=IF(E{row}=0,0,F{row}/E{row})')
        cell.font = Font(name=FONT)
        cell.number_format = "0.0%"
    for col, w in zip("ABCDEFG", [16, 10, 3, 8, 10, 12, 8]):
        wo.column_dimensions[col].width = w

    # ---- Aciklama
    wl = wb.create_sheet("Aciklama")
    wl.sheet_view.showGridLines = False
    lines = [
        ("Takip Cizelgesi — kullanim", True, 13),
        ("", False, 10),
        ("Bu cizelge backlog'dan uretildi. Backlog kaynaktir; ikisini senkron tut.", False, 10),
        ("", False, 10),
        ("Elle doldurulan sutunlar:", True, 11),
        ("- Durum: acilir listeden sec (Yapilacak / Devam / Bloke / Tamamlandi)", False, 10),
        ("- Sorumlu, Baslangic, Bitis, Not", False, 10),
        ("", False, 10),
        ("Uretilen sutunlar (elle degistirme, backlog'u degistir):", True, 11),
        ("- ID, Faz, Epik, Gorev, Katman, Tahmin, Bagimlilik, Hakem", False, 10),
        ("", False, 10),
        ("Hakem sutunu:", True, 11),
        ("- Kabul kriterindeki '**Hakem:**' satirindan gelir; bos olmasi normaldir.", False, 10),
        ("- Bos = o gorevin 'Tamamlandi'si oz-beyandir (kimse disaridan dogrulamadi).", False, 10),
        ("- Rapor bunu ayri bir gosterge olarak basar; ilerleme yuzdesini DEGISTIRMEZ.", False, 10),
        ("", False, 10),
        ("Ozet sekmesi formullerle hesaplanir; elle sayi girme.", False, 10),
        ("Rapor icin: python progress.py", False, 10),
    ]
    for i, (txt, bold, size) in enumerate(lines, start=1):
        wl.cell(row=i, column=1, value=txt).font = Font(
            name=FONT, bold=bold, size=size,
            color="1F3864" if bold and size >= 13 else "000000")
    wl.column_dimensions["A"].width = 95

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--backlog", default="03-gorev-listesi.md")
    ap.add_argument("--out", default="tracker.xlsx")
    ap.add_argument("--config", default=None,
                    help="iskele.config.json (faz sirasi icin; yoksa ID'den turetilir)")
    a = ap.parse_args()

    if not Path(a.backlog).exists():
        sys.exit(f"HATA: backlog bulunamadi: {a.backlog}")

    tasks, problems = parse_backlog(a.backlog)
    for p in problems:
        print(f"  UYARI: {p}", file=sys.stderr)
    if not tasks:
        sys.exit(2)

    if a.config and Path(a.config).exists():
        phases = json.loads(Path(a.config).read_text(encoding="utf-8")).get("phases")
    else:
        phases = None
    if not phases:
        seen = []
        for t in tasks:
            if t["faz"] and t["faz"] not in seen:
                seen.append(t["faz"])
        phases = sorted(seen)

    unknown = sorted({t["faz"] for t in tasks} - set(phases))
    if unknown:
        print(f"  UYARI: yapilandirmada olmayan faz kodlari: {unknown}", file=sys.stderr)

    build(tasks, a.out, phases)
    print(f"{len(tasks)} gorev · fazlar: {', '.join(phases)} -> {a.out}")


if __name__ == "__main__":
    main()
