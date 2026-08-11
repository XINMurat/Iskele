#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
progress.py — efor-agirlikli ilerleme ureteci (iskele).

tracker.xlsx (Takip sekmesi) + iskele.config.json  ->  raporun GEN bolgelerini yazar.

Ilke:
  * SAYILAR veriden gelir (xlsx).
  * PROSE yapilandirmadan gelir (iskele.config.json).
  * GEN:...:BEGIN/END disindaki her seye DOKUNULMAZ (elle duzenlenebilir kalir).
  * Bilinmeyen deger SESSIZCE varsayilana dusmez: uyarir ya da yazmaz.

Kullanim:
    python progress.py                       # varsayilan yollar
    python progress.py --check               # yazmadan ozet bas
    python progress.py --xlsx X --html Y --config C
    python progress.py --self-test           # gomulu regresyon testleri

Cikis kodlari: 0 basarili · 2 dogrulama hatasi · 3 degismez ihlali.
Bagimlilik: openpyxl
"""
import argparse, json, re, subprocess, sys, tempfile
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

COL = dict(id="ID", faz="Faz", epik="Epik", tahmin="Tahmin", durum="Durum")
# Opsiyonel sutun: eski cizelgelerde yok. YOKSA gosterge hesaplanmaz ve rapora
# "veri yok" yazilir — %0 yazilmaz. Veri yoklugunu sifir olarak gostermek,
# "sessiz varsayim" kirmizi cizgisinin ta kendisidir.
OPT_COL = dict(hakem="Hakem")

DEFAULTS = {
    "phases": [],
    "effort_weights": {"S": 0.75, "M": 1.5, "L": 4.0},
    "status_credit": {"Tamamlandi": 1.0, "Devam": 0.0,
                      "Bloke": 0.0, "Yapilacak": 0.0},
    "workdays_per_month": 21,
    "phase_meta": {}, "epic_display": {}, "static_steps": [], "flow": [],
}

# Turkce-duyarli katlama: xlsx "Tamamlandi" tutar, insan eli "Tamamlandı" yazar.
_TR_FOLD = str.maketrans("ıİşŞğĞüÜöÖçÇ", "iIsSgGuUoOcC")


def norm(s):
    return str(s or "").strip().translate(_TR_FOLD).lower()


def esc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def fmt(x):
    return ("%.2f" % x).rstrip("0").rstrip(".")


def git_short():
    try:
        return subprocess.check_output(["git", "rev-parse", "--short", "HEAD"],
                                       stderr=subprocess.DEVNULL, text=True).strip() or "—"
    except Exception:
        return "—"


def load_config(path):
    cfg = dict(DEFAULTS)
    if path and Path(path).exists():
        cfg.update(json.loads(Path(path).read_text(encoding="utf-8")))
    return cfg


# ---------------------------------------------------------------- VERI
def load_tasks(xlsx, cfg):
    """xlsx -> (tasks, issues). issues: ('ERROR'|'WARN', mesaj)

    Opsiyonel sutunlarin VARLIGI her gorevin `_cols` alaninda tasinir; bos
    hucre ile hic olmayan sutun ayri seylerdir ve rapor ikisini ayri basar.
    """
    wb = load_workbook(xlsx, data_only=True)
    if "Takip" not in wb.sheetnames:
        sys.exit("HATA: 'Takip' sekmesi yok.")
    ws = wb["Takip"]
    headers = {c.value: i for i, c in enumerate(ws[1])}
    for name in COL.values():
        if name not in headers:
            sys.exit(f"HATA: '{name}' sutunu Takip sekmesinde yok.")

    phases = cfg["phases"]
    st_alias = {norm(k): k for k in cfg["status_credit"]}
    st_alias.update({norm("Tamamlandı"): "Tamamlandi",
                     norm("Yapılacak"): "Yapilacak"})
    ef_alias = {norm(k): k for k in cfg["effort_weights"]}
    ph_alias = {norm(p): p for p in phases}

    tasks, issues, seen = [], [], {}
    for rno, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tid = row[headers[COL["id"]]]
        if not tid:
            continue
        tid = str(tid).strip()
        if tid in seen:
            issues.append(("ERROR", f"satir {rno}: '{tid}' ID'si mukerrer "
                                    f"(ilk: satir {seen[tid]})"))
        seen[tid] = rno

        raw = {k: str(row[headers[v]] or "").strip() for k, v in COL.items()}
        raw.update({k: (str(row[headers[v]] or "").strip() if v in headers else "")
                    for k, v in OPT_COL.items()})

        faz = ph_alias.get(norm(raw["faz"]))
        if faz is None:
            issues.append(("ERROR", f"satir {rno} [{tid}]: gecersiz Faz "
                                    f"'{raw['faz']}' (beklenen: {'/'.join(phases)})"))
            faz = raw["faz"]

        tahmin = ef_alias.get(norm(raw["tahmin"]))
        if tahmin is None:
            issues.append(("ERROR", f"satir {rno} [{tid}]: gecersiz Tahmin "
                                    f"'{raw['tahmin']}' "
                                    f"(beklenen: {'/'.join(cfg['effort_weights'])})"))
            tahmin = raw["tahmin"]

        if not raw["durum"]:
            durum = "Yapilacak"
            issues.append(("WARN", f"satir {rno} [{tid}]: Durum bos -> 'Yapilacak'"))
        else:
            durum = st_alias.get(norm(raw["durum"]))
            if durum is None:
                issues.append(("ERROR", f"satir {rno} [{tid}]: gecersiz Durum "
                                        f"'{raw['durum']}'"))
                durum = raw["durum"]
            elif durum != raw["durum"]:
                issues.append(("WARN", f"satir {rno} [{tid}]: Durum "
                                       f"'{raw['durum']}' -> '{durum}' normalize edildi"))

        if not raw["epik"]:
            issues.append(("ERROR", f"satir {rno} [{tid}]: Epik bos"))

        tasks.append(dict(id=tid, faz=faz, epik=raw["epik"],
                          tahmin=tahmin, durum=durum, hakem=raw["hakem"],
                          _cols={k: (v in headers) for k, v in OPT_COL.items()}))

    if not tasks:
        issues.append(("ERROR", "Takip sekmesinde hic gorev satiri yok"))
    return tasks, issues


def epic_code(epik):
    parts = str(epik or "").split()
    return parts[0] if parts else "(epik yok)"


def compute(tasks, cfg):
    W, CR, phases = cfg["effort_weights"], cfg["status_credit"], cfg["phases"]
    epics = {}
    ph = {p: dict(eff=0.0, done=0.0, n=0, dn=0) for p in phases}
    for t in tasks:
        e = W.get(t["tahmin"], 0.0)
        c = CR.get(t["durum"], 0.0)
        ep = epics.setdefault(epic_code(t["epik"]),
                              dict(eff=0.0, done=0.0, n=0, dn=0))
        ep["eff"] += e; ep["done"] += e * c; ep["n"] += 1; ep["dn"] += (1 if c >= 1 else 0)
        if t["faz"] in ph:
            d = ph[t["faz"]]
            d["eff"] += e; d["done"] += e * c; d["n"] += 1; d["dn"] += (1 if c >= 1 else 0)

    total_eff = sum(p["eff"] for p in ph.values())
    epic_eff = sum(e["eff"] for e in epics.values())
    if abs(epic_eff - total_eff) > 1e-9:
        orphan = sorted({t["faz"] for t in tasks if t["faz"] not in ph})
        raise ValueError(
            f"DEGISMEZ IHLALI: epik efor toplami {epic_eff:g} != faz toplami "
            f"{total_eff:g}. Faz sutunu {phases} disinda deger iceriyor: "
            f"{orphan or '?'} — bu gorevler cubukta gorunur ama toplama girmez.")

    # --- ikinci gosterge: tamamlanan eforun ne kadari oz-beyan? -------------
    # Ilerleme yuzdesi "bitti denen is"i olcer, "dogrulanmis is"i degil. Tek
    # gostergeli rapor, gostergenin olcmedigi seyi sifir degil GORUNMEZ yapar.
    # Burada olculen sey mutevazi ve dogru: tamamlanmis eforun ne kadarinin
    # kabul kriterinde YAZARDAN BASKA bir hakem adi var. Kriterin fiilen
    # kosuldugunu iddia ETMEZ — onu bu script bilemez.
    has_col = bool(tasks) and tasks[0].get("_cols", {}).get("hakem", False)
    arb_done = sum(W.get(t["tahmin"], 0.0) * CR.get(t["durum"], 0.0)
                   for t in tasks if str(t.get("hakem", "")).strip())
    arb_n = sum(1 for t in tasks
                if str(t.get("hakem", "")).strip() and CR.get(t["durum"], 0.0) >= 1)
    total_done = sum(p["done"] for p in ph.values())
    total_dn = sum(p["dn"] for p in ph.values())

    return dict(epics=epics, phases=ph, total_eff=total_eff,
                arbiter_col=has_col,
                arbiter_done=arb_done, arbiter_dn=arb_n,
                arbiter_pct=(round(100 * arb_done / total_done)
                             if has_col and total_done else None),
                total_done=total_done,
                total_n=sum(p["n"] for p in ph.values()),
                total_dn=total_dn,
                overall_pct=round(100 * total_done / total_eff) if total_eff else 0)


def active_phase(C, cfg):
    for p in cfg["phases"]:
        d = C["phases"][p]
        if d["eff"] and d["done"] < d["eff"] - 1e-9:
            return p
    return None


# ---------------------------------------------------------------- RENDER
def r_chips(C, cfg):
    ap = active_phase(C, cfg)
    nxt = "tamam"
    if ap:
        meta = cfg["phase_meta"].get(ap, {})
        nxt = f"{meta.get('ms', ap)} — {meta.get('title', ap).split('· ')[-1]}"
    return ('<div class="meta">\n'
            f'      <span class="chip"><span class="dot d-todo"></span>Program '
            f'<b>%{C["overall_pct"]}</b> · {C["total_dn"]}/{C["total_n"]} gorev · '
            f'~{fmt(C["total_eff"])} gun</span>\n'
            f'      <span class="chip"><span class="dot '
            f'{"d-prog" if ap else "d-done"}"></span>Siradaki <b>{esc(nxt)}</b></span>\n'
            f'      <span class="chip">git {git_short()}</span>\n'
            f'      <span class="chip">{date.today().isoformat()}</span>\n'
            '    </div>')


def r_kpi(C, cfg):
    ph, phases = C["phases"], cfg["phases"]
    wpm = cfg["workdays_per_month"] or 21
    out = ['<div class="kpis">']
    cum = 0.0
    for p in phases[:-1]:
        cum += ph[p]["eff"]
        out.append(f'      <div class="kpi"><div class="v num">~{fmt(cum)}</div>'
                   f'<div class="l">{esc(p)} sonuna kadar</div>'
                   f'<div class="l2">~{fmt(round(cum / wpm, 1))} ay</div></div>')
    out.append(f'      <div class="kpi"><div class="v num">~{fmt(C["total_eff"])}</div>'
               f'<div class="l">Tum program</div>'
               f'<div class="l2">~{fmt(round(C["total_eff"] / wpm, 1))} ay</div></div>')
    rem = fmt(C["total_eff"] - C["total_done"])
    out.append('      <div class="kpi" style="border-color:color-mix(in srgb,'
               'var(--accent) 40%,var(--line))">'
               f'<div class="v num">{C["overall_pct"]}<small>%</small></div>'
               f'<div class="l">Program geneli (olculen)</div>'
               f'<div class="l2">{C["total_dn"]}/{C["total_n"]} gorev · '
               f'<b>~{rem} gun kalan</b></div></div>')
    out.append('    </div>')
    return "\n".join(out)


def r_cards(C, cfg):
    ap = active_phase(C, cfg)
    out = ['<div class="cards">']
    for p in cfg["phases"]:
        d = C["phases"][p]
        meta = cfg["phase_meta"].get(p, {})
        pct = round(100 * d["done"] / d["eff"]) if d["eff"] else 0
        if pct >= 100:
            edge, badge = "done", '<span class="status s-done">tamam</span>'
        elif p == ap:
            label = f'siradaki ({meta.get("ms", p)})' if pct == 0 else f'devam (%{pct})'
            edge, badge = "prog", f'<span class="status s-prog">{label}</span>'
        else:
            dep = meta.get("dep") or meta.get("ms", p)
            edge, badge = "todo", f'<span class="status s-wait">{esc(dep)} kapisina bagli</span>'
        out.append(
            f'      <div class="card" style="--edge:var(--{edge})">\n'
            f'        <h2>{esc(meta.get("title", p))} {badge}</h2>\n'
            f'        <div class="repo">{esc(meta.get("repo", ""))}</div>\n'
            f'        <p>{esc(meta.get("desc", ""))}</p>\n'
            f'        <div class="row"><span>Ilerleme</span><b class="num">{pct}%</b></div>\n'
            f'        <span class="track" style="height:6px;margin:6px 0 10px">'
            f'<span class="fill" style="width:{pct}%"></span></span>\n'
            f'        <div class="row"><span>Efor</span>'
            f'<b class="num">~{fmt(d["eff"])} gun</b></div>\n'
            f'        <div class="row"><span>Kalan</span>'
            f'<b class="num">~{fmt(d["eff"] - d["done"])} gun</b></div>\n'
            f'        <div class="row"><span>Gorev</span>'
            f'<b class="num">{d["dn"]}/{d["n"]}</b></div>\n'
            f'        <div class="row"><span>Yigin</span>'
            f'<b>{esc(meta.get("stack", "—"))}</b></div>\n'
            f'      </div>')
    out.append('    </div>')
    return "\n".join(out)


def r_bars(C, cfg):
    items = sorted(C["epics"].items(), key=lambda kv: (-kv[1]["eff"], kv[0]))
    max_eff = max((v["eff"] for _, v in items), default=1) or 1
    out = ['<div class="bars">']
    for code, d in items:
        name = cfg["epic_display"].get(code, code)
        pct = round(100 * d["done"] / d["eff"]) if d["eff"] else 0
        out.append(f'      <div class="bar"><span class="name">{esc(name)}</span>'
                   f'<span class="lane"><span class="track" '
                   f'style="width:{round(d["eff"] / max_eff * 100)}%">'
                   f'<span class="fill" style="width:{pct}%"></span></span></span>'
                   f'<span class="pct">{pct}%</span>'
                   f'<span class="ka">{fmt(d["eff"])}</span>'
                   f'<span class="rem">{fmt(d["eff"] - d["done"])} kalan</span></div>')
    out.append(f'      <div class="bar total"><span class="name">Toplam '
               f'({len(items)} epik)</span><span class="lane"></span>'
               f'<span class="pct">{C["overall_pct"]}%</span>'
               f'<span class="ka">{fmt(C["total_eff"])}</span>'
               f'<span class="rem">{fmt(C["total_eff"] - C["total_done"])} kalan</span></div>')
    out.append('    </div>')
    return "\n".join(out)


def r_timeline(C, cfg):
    ap = active_phase(C, cfg)
    out = ['<div class="tl">']
    for s in cfg["static_steps"]:
        out.append(f'      <div class="step ok"><span class="mk">{esc(s.get("mk", "✓"))}</span>'
                   f'<div><div class="t">{esc(s.get("t", ""))}</div>'
                   f'<div class="m">{esc(s.get("m", ""))}</div></div>'
                   f'<span class="when">tamam</span></div>')
    for step in cfg["flow"]:
        p = step.get("phase")
        d = C["phases"].get(p, dict(eff=0, done=0))
        pct = round(100 * d["done"] / d["eff"]) if d["eff"] else 0
        if step.get("kind") == "gate":
            cls, when = ("step gate ok", "acik") if pct >= 100 else ("step gate", "karar")
        elif pct >= 100:
            cls, when = "step ok", "tamam"
        elif pct > 0:
            cls, when = "step on", "devam"
        elif p == ap:
            cls, when = "step on", "siradaki"
        else:
            cls, when = "step", "—"
        out.append(f'      <div class="{cls}"><span class="mk">{esc(step.get("mk", "▸"))}</span>'
                   f'<div><div class="t">{esc(step.get("t", ""))}</div>'
                   f'<div class="m">{esc(step.get("m", ""))}</div></div>'
                   f'<span class="when">{when}</span></div>')
    out.append('    </div>')
    return "\n".join(out)


def r_hakem(C, cfg):
    """Tamamlanan eforun ne kadari disaridan hakemli — ilerlemenin yaninda,
    ILERLEMENIN ICINDE degil. Bu sayiyi yuzdeye katmak, iki ayri seyi
    (bitti / dogrulandi) tek gostergeye eritirdi."""
    if not C["arbiter_col"]:
        return ('<div class="note">\n'
                '      <b>Hakem gostergesi: veri yok.</b> Cizelgede <code>Hakem</code>\n'
                '      sutunu bulunmuyor (eski surum). Bu <em>%0</em> demek degildir —\n'
                '      olculmedi demektir. Cizelgeyi <code>backlog_to_tracker.py</code>\n'
                '      ile yeniden uret.\n'
                '    </div>')
    if C["total_done"] <= 0:
        return ('<div class="note">\n'
                '      <b>Hakem gostergesi: henuz tamamlanan is yok.</b> Oran, ilk\n'
                '      gorev kapandiginda anlam kazanir.\n'
                '    </div>')

    pct = C["arbiter_pct"]
    self_rep = 100 - pct
    return ('<div class="note">\n'
            f'      <div class="row"><span>Tamamlanan eforun hakemli olani</span>'
            f'<b class="num">%{pct}</b></div>\n'
            f'      <span class="track" style="height:6px;margin:6px 0 10px">'
            f'<span class="fill" style="width:{pct}%"></span></span>\n'
            f'      <div class="row"><span>Oz-beyanda kalan</span>'
            f'<b class="num">%{self_rep}</b></div>\n'
            f'      <div class="row"><span>Hakemli kapanan gorev</span>'
            f'<b class="num">{C["arbiter_dn"]}/{C["total_dn"]}</b></div>\n'
            '      <p class="l2">Olculen sey: kabul kriterinde yazardan baska bir\n'
            '      hakem adi gecen gorevlerin, tamamlanan efor icindeki payi.\n'
            '      Kriterin fiilen kosuldugunu <b>gostermez</b> — bunu hicbir\n'
            '      cizelge bilemez; Tamamlandi Tanimi bilir. Dusuk oran, isin\n'
            '      kotu oldugunu degil, "bitti" hukmunun buyuk olcude isi yapanin\n'
            '      kendi beyanina dayandigini soyler.</p>\n'
            '    </div>')


RENDERERS = dict(CHIPS=r_chips, KPI=r_kpi, CARDS=r_cards,
                 BARS=r_bars, TIMELINE=r_timeline, HAKEM=r_hakem)


def patch(html, key, block):
    pat = re.compile(r'(<!-- GEN:%s:BEGIN[^>]*-->\n    ).*?(\n<!-- GEN:%s:END -->)'
                     % (key, key), re.S)
    if not pat.search(html):
        return html, False
    return pat.sub(lambda m: m.group(1) + block + m.group(2), html), True


# ---------------------------------------------------------------- SELF-TEST
def self_test():
    """Gomulu regresyon: mutlu yol + uc sessiz-hata vakasi."""
    from openpyxl import Workbook
    cfg = dict(DEFAULTS, phases=["F0", "F1"])
    tmp = Path(tempfile.mkdtemp())
    ok = True

    def make(rows):
        wb = Workbook(); ws = wb.active; ws.title = "Takip"
        ws.append(["ID", "Faz", "Epik", "Tahmin", "Durum"])
        for r in rows:
            ws.append(r)
        p = tmp / "t.xlsx"; wb.save(p); return p

    def check(label, cond):
        nonlocal ok
        print(f"  {'PASS' if cond else 'FAIL'}  {label}")
        ok = ok and cond

    # 1) mutlu yol: S+L tamam, M yapilacak -> (0.75+4)/(0.75+4+1.5)=76%
    t, iss = load_tasks(make([
        ["A-1", "F0", "F0.1 Bir", "S", "Tamamlandi"],
        ["A-2", "F0", "F0.1 Bir", "L", "Tamamlandi"],
        ["B-1", "F1", "F1.1 Iki", "M", "Yapilacak"]]), cfg)
    C = compute(t, cfg)
    check("mutlu yol: hata yok", not [m for l, m in iss if l == "ERROR"])
    check("efor toplami 6.25", abs(C["total_eff"] - 6.25) < 1e-9)
    check("ilerleme %76", C["overall_pct"] == 76)
    check("aktif faz F1", active_phase(C, cfg) == "F1")

    # 2) gecersiz tahmin -> ERROR (sessizce M olmaz)
    _, iss = load_tasks(make([["A-1", "F0", "F0.1 Bir", "XL", "Yapilacak"]]), cfg)
    check("gecersiz Tahmin ERROR uretir",
          any(l == "ERROR" and "Tahmin" in m for l, m in iss))

    # 3) Turkce-i 'Tamamlandı' -> normalize + kredi alir
    t, iss = load_tasks(make([["A-1", "F0", "F0.1 Bir", "M", "Tamamlandı"]]), cfg)
    C = compute(t, cfg)
    check("Turkce-i Durum normalize edilir",
          any(l == "WARN" and "normalize" in m for l, m in iss))
    check("Turkce-i Durum kredi alir (%100)", C["overall_pct"] == 100)

    # 4) faz listesi disi -> degismez ihlali
    t, iss = load_tasks(make([
        ["A-1", "F0", "F0.1 Bir", "M", "Yapilacak"],
        ["Z-1", "F9", "F9.9 Hayalet", "L", "Tamamlandi"]]), cfg)
    try:
        compute(t, cfg)
        check("faz disi kayit degismez ihlali uretir", False)
    except ValueError:
        check("faz disi kayit degismez ihlali uretir", True)

    # 5) Hakem sutunu YOKSA gosterge %0 degil "olculmedi" olmali
    t, _ = load_tasks(make([["A-1", "F0", "F0.1 Bir", "M", "Tamamlandi"]]), cfg)
    C = compute(t, cfg)
    check("Hakem sutunu yoksa gosterge None (0 degil)",
          C["arbiter_col"] is False and C["arbiter_pct"] is None)

    # 6) Hakem sutunu varsa oran TAMAMLANAN efor uzerinden hesaplanir.
    #    L(4) hakemli tamam + S(0.75) hakemsiz tamam + M(1.5) hakemli ama
    #    yapilacak -> 4/4.75 = %84. Bitmemis is orana girmemeli: girseydi
    #    "hakem yazdim" demek, dogrulanmis is gibi gorunurdu.
    def make_h(rows):
        wb = Workbook(); ws = wb.active; ws.title = "Takip"
        ws.append(["ID", "Faz", "Epik", "Tahmin", "Durum", "Hakem"])
        for r in rows:
            ws.append(r)
        p = tmp / "th.xlsx"; wb.save(p); return p

    t, _ = load_tasks(make_h([
        ["A-1", "F0", "F0.1 Bir", "L", "Tamamlandi", "pytest tests/a.py"],
        ["A-2", "F0", "F0.1 Bir", "S", "Tamamlandi", ""],
        ["B-1", "F1", "F1.1 Iki", "M", "Yapilacak", "pytest tests/b.py"]]), cfg)
    C = compute(t, cfg)
    check("hakemli oran tamamlanan efor uzerinden (%84)", C["arbiter_pct"] == 84)
    check("hakemli kapanan gorev sayisi 1", C["arbiter_dn"] == 1)
    check("hakem orani ilerleme yuzdesini degistirmez", C["overall_pct"] == 76)

    print("SELF-TEST:", "BASARILI" if ok else "BASARISIZ")
    return 0 if ok else 1


# ---------------------------------------------------------------- MAIN
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="tracker.xlsx")
    ap.add_argument("--html", default="07-ilerleme-raporu.html")
    ap.add_argument("--config", default="iskele.config.json")
    ap.add_argument("--check", action="store_true", help="yazma, ozet bas")
    ap.add_argument("--force", action="store_true", help="hatalara ragmen yaz (onerilmez)")
    ap.add_argument("--self-test", action="store_true", help="gomulu regresyon testleri")
    a = ap.parse_args()

    if a.self_test:
        sys.exit(self_test())

    cfg = load_config(a.config)
    if not cfg["phases"]:
        sys.exit("HATA: yapilandirmada 'phases' bos "
                 "(iskele.config.json olustur; ornek: assets/iskele.config.example.json)")

    tasks, issues = load_tasks(a.xlsx, cfg)
    errors = [m for l, m in issues if l == "ERROR"]
    for l, m in issues:
        print(f"  {'HATA ' if l == 'ERROR' else 'UYARI'}: {m}", file=sys.stderr)
    if errors and not a.force:
        print(f"\n{len(errors)} dogrulama hatasi — rapor YAZILMADI. "
              f"Cizelgeyi duzelt ya da --force kullan.", file=sys.stderr)
        sys.exit(2)

    try:
        C = compute(tasks, cfg)
    except ValueError as e:
        print(f"\n{e}", file=sys.stderr)
        sys.exit(3)

    print(f"Gorev: {C['total_dn']}/{C['total_n']} · "
          f"Efor: {fmt(C['total_done'])}/{fmt(C['total_eff'])} gun · "
          f"Program %{C['overall_pct']}")
    for p in cfg["phases"]:
        d = C["phases"][p]
        pct = round(100 * d["done"] / d["eff"]) if d["eff"] else 0
        print(f"  {p}: %{pct:>3} · {d['dn']}/{d['n']} gorev · {fmt(d['eff'])} gun")

    if not C["arbiter_col"]:
        print("  Hakem: olculmedi (cizelgede 'Hakem' sutunu yok) — %0 DEGIL")
    elif C["total_done"] <= 0:
        print("  Hakem: henuz tamamlanan is yok")
    else:
        print(f"  Hakem: tamamlanan eforun %{C['arbiter_pct']}'i hakemli · "
              f"%{100 - C['arbiter_pct']}'i oz-beyan "
              f"({C['arbiter_dn']}/{C['total_dn']} gorev)")

    if a.check:
        return
    if not Path(a.html).exists():
        sys.exit(f"HATA: rapor bulunamadi: {a.html}")

    html = Path(a.html).read_text(encoding="utf-8")
    applied, missing = [], []
    for key, fn in RENDERERS.items():
        html, done = patch(html, key, fn(C, cfg))
        (applied if done else missing).append(key)
    Path(a.html).write_text(html, encoding="utf-8")
    print(f"Guncellendi: {a.html} (bolgeler: {', '.join(applied) or 'yok'})")
    if missing:
        print(f"  UYARI: raporda bulunmayan GEN bolgeleri: {', '.join(missing)}",
              file=sys.stderr)


if __name__ == "__main__":
    main()
